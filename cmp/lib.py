import copy
from random import randint
import os
import django
import numpy

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "dipl.settings")
django.setup()

import numpy as np
from openpyxl import load_workbook

from cmp.models import Weights, GroupWeights


file = './data.xlsx'

wb = load_workbook(filename=file, data_only=True)

ws = wb.get_sheet_by_name('Входные данные')


def cmp(weights, param='kfavg', group=-1):
    ws = wb.get_sheet_by_name('Входные данные')
    data = [list(map(lambda x: x.value, row)) for row in ws['D2':'AV16']]
    if param == 'kf':
        ws = wb.get_sheet_by_name('Входные данные (кф)')
        data = [list(map(lambda x: x.value, row)) for row in ws['D2':'AV16']]
    elif param == 'avg' or param == 'c':
        ws = wb.get_sheet_by_name('Входные данные (кол)')
        data = [list(map(lambda x: x.value, row)) for row in ws['D2':'AV16']]
        if param == 'avg':
            counts = [cell[0].value for cell in ws['AW2':'AW16']]
            for i in range(len(data)):
                for j in range(len(data[i])):
                    data[i][j] = data[i][j] / counts[i]

    name_groups = [cell[0].value for cell in ws['A2':'A16']]
    old_ranking = [cell[0].value for cell in ws['B2':'B16']]
    academic_performance = [cell[0].value for cell in ws['C2':'C16']]

    if group != -1:
        data.pop(group)
        name_groups.pop(group)
        old_ranking.pop(group)
        academic_performance.pop(group)

    new_ranking = []

    culture = []
    sport = []
    science = []
    social_work = []
    patriotism = []

    for group in data:
        culture.append(sum([sc * w for sc, w in zip(group[:5], weights[:5])]))
        sport.append(sum([sc * w for sc, w in zip(group[5:10], weights[5:10])]))
        patriotism.append(sum([sc * w for sc, w in zip(group[10:13], weights[10:13])]))
        science.append(sum([sc * w for sc, w in zip(group[13:40], weights[13:40])]))
        social_work.append(sum([sc * w for sc, w in zip(group[40:45], weights[40:45])]))
    max_culture, min_culture = max(culture), min(culture)
    max_sport, min_sport = max(sport), min(sport)
    max_science, min_science = max(science), min(science)
    max_social_work, min_social_work = max(social_work), min(social_work)
    max_patriotism, min_patriotism = max(patriotism), min(patriotism)
    for i in range(len(data)):
        culture_norm = (culture[i] - min_culture) / (max_culture - min_culture) * 10
        sport_norm = (sport[i] - min_sport) / (max_sport - min_sport) * 10
        science_norm = (science[i] - min_science) / (max_science - min_science) * 10
        social_norm = (social_work[i] - min_social_work) / (max_social_work - min_social_work) * 10
        patriotism_norm = (patriotism[i] - min_patriotism) / (max_patriotism - min_patriotism) * 10
        new_ranking.append(culture_norm + sport_norm + science_norm + social_norm + patriotism_norm +
                           academic_performance[
                               i])
    return old_ranking, new_ranking, [abs(old - new) for old, new in zip(old_ranking, new_ranking)], name_groups


pattern = [["Пост", 0], ["Д1", 0], ["Д2", 0], ["Д3", 0], ["уч", 0]]
olimp_pattern = [["Д1", 0], ["Д2", 0], ["Д3", 0], ["уч", 0]]
patriotism_ = [["Пост", 0], ["грам/дип", 0], ["уч", 0]]
science = {"Публикации": [["У1", 0], ["У2", 0]],
           "Школьный": [["Д1", 0], ["Д2", 0], ["Д3", 0], ["уч", 0]],
           "Муниципальный": [["Д1", 0], ["Д2", 0], ["Д3", 0], ["уч", 0]],
           "Региональный": [["Д1", 0], ["Д2", 0], ["Д3", 0], ["уч", 0]],
           "Всероссийский": [["Д1", 0], ["Д2", 0], ["Д3", 0], ["уч", 0]],
           "Общие": [["Д1", 0], ["Д2", 0], ["Д3", 0], ["уч", 0]],
           "Конференции": [["Д1", 0], ["Д2", 0], ["Д3", 0], ["уч", 0]],
           "Похвальные грамоты/письма": [["парам", 0]]}


def fill_inputs(weights):
    culture, social_work, sport, patriotism = np.array(copy.deepcopy(pattern)), np.array(
        copy.deepcopy(pattern)), np.array(
        copy.deepcopy(pattern)), np.array(copy.deepcopy(patriotism_))
    culture[:, 1], sport[:, 1], social_work[:, 1], patriotism[:, 1] = weights[:5], weights[5:10], weights[
                                                                                                  40:45], weights[10:13]
    articles = np.array(copy.deepcopy(science["Публикации"]))
    articles[:, 1] = weights[13:15]
    olymp_conf = {}
    c = 15
    for key in list(science.keys())[1:-1]:
        olymp_conf[key] = np.array(copy.deepcopy(olimp_pattern))
        olymp_conf[key][:, 1] = weights[c:c + 4]
        olymp_conf[key] = olymp_conf[key].tolist()
        c += 4
    certs = copy.deepcopy(science["Похвальные грамоты/письма"])
    certs[0][1] = weights[39]
    science_ = {}
    science_["Публикации"] = articles.tolist()
    science_ = {**science_, **olymp_conf}
    science_["Похвальные грамоты/письма"] = certs

    all_criteria = [{'Культура': culture.tolist()}, {"Спорт": sport.tolist()}, {"Патриотизм": patriotism.tolist()},
                    science_,
                    {"Общественная": social_work.tolist()}
                    ]

    return all_criteria


K = 20


def put_best(weights, sum, param='kfavg'):
    objects = Weights.objects.filter(type=param).order_by('-deviations_sum')
    if objects.count() < K:
        Weights.objects.create(weights=' '.join(str(e) for e in weights), deviations_sum=sum, type=param)
    elif objects.first().deviations_sum > sum:
        objects.first().delete()
        Weights.objects.create(weights=' '.join(str(e) for e in weights), deviations_sum=sum, type=param)


def generate_random_weights(iter_count):
    random_weights = [[randint(1, 9) for i in range(45)] for i in range(iter_count)]
    for weights in random_weights:
        put_best(weights, sum(cmp(weights, param='c')[2]), param='c')


name_groups = [cell[0].value for cell in ws['A2':'A16']]
old_ranking = [cell[0].value for cell in ws['B2':'B16']]
academic_performance = [cell[0].value for cell in ws['C2':'C16']]


def generate_random_group_weights(iter_count, group_number, param):
    random_weights = [[randint(1, 9) for i in range(45)] for i in range(iter_count)]
    min = sum(cmp(random_weights[0], param=param, group=group_number)[2])
    best_weights = random_weights[0]
    random_weights.pop(0)
    for weights in random_weights:
        result = sum(cmp(weights, param=param, group=group_number)[2])
        if result < min:
            min = result
            best_weights = weights
    GroupWeights.objects.create(group_name=name_groups[group_number], weights=' '.join(str(e) for e in best_weights),
                                deviations_sum=min, type=param)


def groups(iter_count):
    for param in ['c', 'avg']:
        for j in range(len(name_groups)):
            generate_random_group_weights(iter_count, j, param)


def each_group_separately(param):
    ws = wb.get_sheet_by_name('Входные данные')
    data = [list(map(lambda x: x.value, row)) for row in ws['D2':'AV16']]
    if param == 'kf':
        ws = wb.get_sheet_by_name('Входные данные (кф)')
        data = [list(map(lambda x: x.value, row)) for row in ws['D2':'AV16']]
    elif param == 'avg' or param == 'c':
        ws = wb.get_sheet_by_name('Входные данные (кол)')
        data = [list(map(lambda x: x.value, row)) for row in ws['D2':'AV16']]
        if param == 'avg':
            counts = [cell[0].value for cell in ws['AW2':'AW16']]
            for i in range(len(data)):
                for j in range(len(data[i])):
                    data[i][j] = data[i][j] / counts[i]
    cmp = []
    func = lambda x: float("{0:.3f}".format(x))
    for i in range(len(name_groups)):
        group_weights = list(map(int, GroupWeights.objects.get(group_name=name_groups[i], type=param).weights.split()))
        new_value = sum([sc * w for sc, w in zip(group_weights, data[i])])
        old_value = old_ranking[i]
        diff = abs(new_value - old_value)
        cmp.append([old_value, func(new_value), func(diff), name_groups[i]])
    return cmp


# print(each_group_separately('kfavg'))

# generate_random_weights(1000000)

# def read_weights():
#     f = open('../static/result/weights2.txt', 'r')
#     for line in f.readlines():
#         Weights.objects.create(weights=line[:89], deviations_sum=float(line[90:].split()[0]), type=line[90:-1].split()[1])
#     f.close()

# def write_weights():
#     objects = Weights.objects.all()
#     f = open('../static/result/weights2.txt', 'w')
#     for object in objects:
#         f.write(object.weights + ' ' + str(object.deviations_sum) + ' ' + object.type + '\n')
#     f.close()

def cal_pop_fitness(equation_inputs, pop):
    # Calculating the fitness value of each solution in the current population.
    # The fitness function calulates the sum of products between each input and its corresponding weight.
    fitness = numpy.sum(pop * equation_inputs, axis=1)
    return fitness


def select_mating_pool(pop, fitness, num_parents):
    # Selecting the best individuals in the current generation as parents for producing the offspring of the next generation.
    parents = numpy.empty((num_parents, pop.shape[1]))
    for parent_num in range(num_parents):
        max_fitness_idx = numpy.where(fitness == numpy.max(fitness))
        max_fitness_idx = max_fitness_idx[0][0]
        parents[parent_num, :] = pop[max_fitness_idx, :]
        fitness[max_fitness_idx] = -99999999999
    return parents


def crossover(parents, offspring_size):
    offspring = numpy.empty(offspring_size)
    # The point at which crossover takes place between two parents. Usually, it is at the center.
    crossover_point = numpy.uint8(offspring_size[1] / 2)

    for k in range(offspring_size[0]):
        # Index of the first parent to mate.
        parent1_idx = k % parents.shape[0]
        # Index of the second parent to mate.
        parent2_idx = (k + 1) % parents.shape[0]
        # The new offspring will have its first half of its genes taken from the first parent.
        offspring[k, 0:crossover_point] = parents[parent1_idx, 0:crossover_point]
        # The new offspring will have its second half of its genes taken from the second parent.
        offspring[k, crossover_point:] = parents[parent2_idx, crossover_point:]
    return offspring


def mutation(offspring_crossover):
    # Mutation changes a single gene in each offspring randomly.
    for idx in range(offspring_crossover.shape[0]):
        # The random value to be added to the gene.
        random_value = numpy.random.uniform(-1.0, 1.0, 1)
        offspring_crossover[idx, 4] = offspring_crossover[idx, 4] + random_value
    return offspring_crossover

