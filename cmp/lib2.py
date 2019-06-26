import pickle
import random

from openpyxl import load_workbook

file = './data3.xlsx'

wb = load_workbook(filename=file, data_only=True)

ws = wb.get_sheet_by_name('Входные данные (кол)')

academic_performance = [cell[0].value for cell in ws['C2':'C9']]

input_file_template = './static/cmp_results/3%s'

criterion_indexes = {
    'culture': (0, 5),
    'sport': (5, 10),
    'patriotism': (10, 13),
    'science': (13, 40),
    'social_work': (40, 45)
}


def save_data(param='avg'):
    file = open(input_file_template % param, 'wb')
    param_values = []
    if param == 'kfavg':
        ws = wb.get_sheet_by_name('Входные данные')
        param_values = [list(map(lambda x: x.value, row)) for row in ws['D2':'AV9']]
    elif param == 'avg' or param == 'c':
        ws = wb.get_sheet_by_name('Входные данные (кол)')
        param_values = [list(map(lambda x: x.value, row)) for row in ws['D2':'AV9']]
        if param == 'avg':
            counts = [cell[0].value for cell in ws['AW2':'AW9']]
            for i in range(len(param_values)):
                for j in range(len(param_values[i])):
                    param_values[i][j] = param_values[i][j] / counts[i]

    ws = wb.get_sheet_by_name('Входные данные (кол)')
    name_groups = [cell[0].value for cell in ws['A2':'A9']]
    jury_results = [cell[0].value for cell in ws['B2':'B9']]
    pickle.dump([param_values, name_groups, jury_results, academic_performance], file)


def calculate_estimations(all_requests_param_values, weights):

    estimations = []
    culture = []
    sport = []
    science = []
    social_work = []
    patriotism = []

    for request_param_values in all_requests_param_values:
        culture.append(sum([sc * w for sc, w in zip(request_param_values[:5], weights[:5])]))
        sport.append(sum([sc * w for sc, w in zip(request_param_values[5:10], weights[5:10])]))
        patriotism.append(sum([sc * w for sc, w in zip(request_param_values[10:13], weights[10:13])]))
        science.append(sum([sc * w for sc, w in zip(request_param_values[13:40], weights[13:40])]))
        social_work.append(sum([sc * w for sc, w in zip(request_param_values[40:45], weights[40:45])]))
    max_culture, min_culture = max(culture), min(culture)
    max_sport, min_sport = max(sport), min(sport)
    max_science, min_science = max(science), min(science)
    max_social_work, min_social_work = max(social_work), min(social_work)
    max_patriotism, min_patriotism = max(patriotism), min(patriotism)
    for i in range(len(all_requests_param_values)):
        culture_norm = (culture[i] - min_culture) / (max_culture - min_culture) * 10
        sport_norm = (sport[i] - min_sport) / (max_sport - min_sport) * 10
        science_norm = (science[i] - min_science) / (max_science - min_science) * 10
        social_norm = (social_work[i] - min_social_work) / (max_social_work - min_social_work) * 10
        patriotism_norm = (patriotism[i] - min_patriotism) / (max_patriotism - min_patriotism) * 10
        estimations.append(culture_norm + sport_norm + science_norm + social_norm + patriotism_norm +
                           academic_performance[i])
    return estimations

